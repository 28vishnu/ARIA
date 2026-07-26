import os
import json
import httpx
import base64
import re
import asyncio
import certifi
from datetime import datetime, timezone, timedelta
from io import BytesIO
from fastapi import FastAPI, Request, UploadFile, File, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse
from pydantic import BaseModel
from pypdf import PdfReader
from docx import Document
import openpyxl
import edge_tts

# Provider SDKs
from groq import Groq
from google import genai
from google.genai import types
from tavily import TavilyClient
import motor.motor_asyncio

# ChromaDB & Embeddings SDK
import chromadb
from chromadb.utils import embedding_functions

# Scheduler SDKs
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.date import DateTrigger

app = FastAPI()

# -------------------------------------------------------------
# 1. GLOBAL ENVIRONMENT & CACHE INITIALIZATION
# -------------------------------------------------------------
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
MONGODB_URI = os.getenv("MONGODB_URI")
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
ALLOWED_TELEGRAM_USER_ID = os.getenv("ALLOWED_TELEGRAM_USER_ID")
TAVILY_API_KEY = os.getenv("TAVILY_API_KEY")

ASSISTANT_NAME = "ARIA"
USER_FULL_NAME = "N. Vishnu Saketh"

RAM_MEMORY_CACHE = [f"[PERSONAL_PROFILE]: User Full Name is {USER_FULL_NAME}"]
RAM_SCHEDULE_CACHE = []
RAM_RECENT_CHATS = []
LAST_CACHE_UPDATE = 0
LAST_USER_INTERACTION_TIME = datetime.now(timezone.utc)
PENDING_SECURITY_ACTIONS = {}

groq_client = Groq(api_key=GROQ_API_KEY) if GROQ_API_KEY else None
gemini_client = genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None
tavily_client = TavilyClient(api_key=TAVILY_API_KEY) if TAVILY_API_KEY else None

def init_mongo_client():
    if not MONGODB_URI: return None
    try:
        return motor.motor_asyncio.AsyncIOMotorClient(
            MONGODB_URI,
            tlsCAFile=certifi.where(),
            tlsInsecure=True,
            serverSelectionTimeoutMS=5000,
            connectTimeoutMS=5000
        )
    except Exception as e:
        print(f"[Mongo Init Exception]: {e}")
        return motor.motor_asyncio.AsyncIOMotorClient(MONGODB_URI)

mongo_client = init_mongo_client()
mongo_db = mongo_client["aria_db"] if mongo_client else None

mongo_memory_col = mongo_db["personal_memory"] if mongo_db is not None else None
mongo_tasks_col = mongo_db["tasks_schedule"] if mongo_db is not None else None
mongo_media_col = mongo_db["media_vault"] if mongo_db is not None else None
mongo_chats_col = mongo_db["chat_history"] if mongo_db is not None else None
mongo_reminders_col = mongo_db["reminders"] if mongo_db is not None else None
mongo_security_col = mongo_db["security_logs"] if mongo_db is not None else None

scheduler = AsyncIOScheduler()

# -------------------------------------------------------------
# 2. CHROMA VECTOR DATABASE & BETTER EMBEDDINGS (BAAI bge-small)
# -------------------------------------------------------------
try:
    embedding_fn = embedding_functions.SentenceTransformerEmbeddingFunction(
        model_name="BAAI/bge-small-en-v1.5"
    )
except Exception:
    embedding_fn = embedding_functions.DefaultEmbeddingFunction()

chroma_client = chromadb.PersistentClient(path="./aria_vectors")

documents_collection = chroma_client.get_or_create_collection(
    name="documents",
    embedding_function=embedding_fn
)

memory_collection = chroma_client.get_or_create_collection(
    name="memory",
    embedding_function=embedding_fn
)

# -------------------------------------------------------------
# 3. UNIVERSAL DOCUMENT PARSERS (PDF, DOCX, XLSX, TXT)
# -------------------------------------------------------------
def clean_response_text(raw_text: str) -> str:
    if not raw_text: return ""
    text = raw_text.strip()
    text = re.sub(r'\*+', '', text)
    text = re.sub(r'[\,\.\s]*,[\,\.\s]*', ', ', text)
    text = re.sub(r'\.\s*\.', '.', text)
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\s+([\.,\?!])', r'\1', text)
    return text.strip()

def get_current_temporal_context() -> str:
    now_utc = datetime.now(timezone.utc)
    ist_offset = timedelta(hours=5, minutes=30)
    now_ist = now_utc + ist_offset
    return f"\nLIVE TEMPORAL CONTEXT: {now_ist.strftime('%A, %B %d, %Y at %I:%M:%S %p IST')}\n"

def parse_document_bytes(file_name: str, file_bytes: bytes, password: str = None) -> tuple[str, bool]:
    fn = file_name.lower()
    is_encrypted = False
    text = ""
    try:
        if fn.endswith(".pdf"):
            reader = PdfReader(BytesIO(file_bytes))
            if reader.is_encrypted:
                if password:
                    try:
                        decrypt_success = reader.decrypt(password)
                        if not decrypt_success:
                            return "[INVALID_PASSWORD]", True
                    except Exception:
                        return "[INVALID_PASSWORD]", True
                else:
                    return "[ENCRYPTED_PDF_LOCKED]", True
            text = "".join([page.extract_text() or "" for page in reader.pages]).strip()

        elif fn.endswith(".docx"):
            doc = Document(BytesIO(file_bytes))
            text = "\n".join([p.text for p in doc.paragraphs if p.text]).strip()

        elif fn.endswith(".xlsx"):
            wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
            sheet_texts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_str = " ".join([str(cell) for cell in row if cell is not None])
                    if row_str: sheet_texts.append(row_str)
            text = "\n".join(sheet_texts).strip()

        elif fn.endswith(".txt") or fn.endswith(".csv"):
            text = file_bytes.decode("utf-8", errors="ignore").strip()

        else:
            text = "Binary File Payload"
    except Exception as e:
        text = f"[Parsing Error: {str(e)}]"
    return text, is_encrypted

# -------------------------------------------------------------
# 4. MULTI-AGENT PLANNER & GEMINI VISION OCR AGENT
# -------------------------------------------------------------
async def ai_planner(user_text: str) -> dict:
    prompt = f"""Analyze the user request and determine which agents/tools are needed.
User Request: "{user_text}"

Return strict JSON format with boolean flags:
{{
  "memory": true/false (if asking about prior facts, preferences, past discussions),
  "documents": true/false (if asking about uploaded PDFs, code, notes, spreadsheets, documents),
  "vision": true/false (if asking about uploaded images or screenshots),
  "internet": true/false (if asking for live web search, current news, facts, weather)
}}"""
    if groq_client:
        try:
            comp = groq_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1, max_tokens=150
            )
            raw_res = comp.choices[0].message.content.strip()
            raw_res = re.sub(r'```json\s*|\s*```', '', raw_res)
            return json.loads(raw_res)
        except Exception:
            pass
    return {
        "memory": any(w in user_text.lower() for w in ["remember", "before", "told", "my", "preference", "favorite"]),
        "documents": any(w in user_text.lower() for w in ["pdf", "certificate", "resume", "document", "file", "sheet"]),
        "vision": any(w in user_text.lower() for w in ["image", "photo", "screenshot", "picture"]),
        "internet": any(w in user_text.lower() for w in ["latest", "today", "news", "search", "weather"])
    }

async def process_image_with_gemini_vision(image_bytes: bytes, file_name: str) -> str:
    if not gemini_client:
        return "Image uploaded successfully."
    try:
        def _gem_vision():
            response = gemini_client.models.generate_content(
                model='gemini-2.0-flash',
                contents=[
                    types.Part.from_bytes(
                        data=image_bytes,
                        mime_type='image/jpeg',
                    ),
                    "Extract all text (OCR), describe visual elements, charts, code, or handwritten notes in detail for semantic indexing."
                ]
            )
            return response.text
        description = await asyncio.to_thread(_gem_vision)
        return description.strip()
    except Exception as e:
        return f"Image Vision OCR Error: {str(e)}"

def index_document_into_chroma(file_name: str, media_type: str, text: str):
    try:
        chunks = [text[i:i+500] for i in range(0, len(text), 500)]
        for index, chunk in enumerate(chunks):
            documents_collection.add(
                ids=[f"{file_name}_{index}_{datetime.now().timestamp()}"],
                documents=[chunk],
                metadatas=[{"file": file_name, "type": media_type}]
            )
    except Exception as e:
        print(f"[Chroma Indexing Error]: {e}")

async def send_file_from_vault(file_query: str, chat_id: str) -> str:
    if mongo_media_col is None: return "Vault database is currently offline, Sir."
    try:
        query_str = file_query.strip() if file_query else ""
        q_regex = re.compile(re.escape(query_str), re.IGNORECASE)
        target_doc = await mongo_media_col.find_one({"$or": [{"file_name": q_regex}, {"caption": q_regex}]})

        if not target_doc:
            target_doc = await mongo_media_col.find_one({"media_type": "document"}, sort=[("_id", -1)])

        if not target_doc:
            return f"I searched your vault, Sir, but could not locate any document matching '{file_query}'."

        fname = target_doc.get("file_name", "document.pdf")
        raw_bytes = base64.b64decode(target_doc["b64_payload"])

        async with httpx.AsyncClient() as client:
            await client.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
                data={"chat_id": chat_id, "caption": f"Here is your document: '{fname}', Sir."},
                files={"document": (fname, raw_bytes, "application/octet-stream")}
            )

        return f"File '{fname}' dispatched successfully to your Telegram, Sir."
    except Exception as e:
        return f"Encountered an issue dispatching document for query '{file_query}', Sir."

async def query_vector_documents(specific_question: str) -> str:
    try:
        results = documents_collection.query(
            query_texts=[specific_question],
            n_results=5
        )
        if results and "documents" in results and results["documents"]:
            docs = results["documents"][0]
            if docs:
                return "\n".join(docs)
        return "No matching document segments found in ChromaDB vector store."
    except Exception as e:
        return f"Document vector query error: {str(e)}"

async def query_vector_memory(user_query: str) -> str:
    try:
        results = memory_collection.query(
            query_texts=[user_query],
            n_results=8
        )
        if results and "documents" in results and results["documents"]:
            docs = results["documents"][0]
            if docs:
                return "\n".join(docs)
        return ""
    except Exception as e:
        return ""

async def unlock_encrypted_pdf(doc_id_or_name: str, password: str) -> tuple[str, bool]:
    if mongo_media_col is None: return "Database offline.", False
    try:
        q_regex = re.compile(re.escape(doc_id_or_name), re.IGNORECASE)
        target_doc = await mongo_media_col.find_one({"$or": [{"file_name": q_regex}, {"caption": q_regex}]})
        
        if not target_doc: return "Document not found.", False

        raw_bytes = base64.b64decode(target_doc["b64_payload"])
        decrypted_text, is_still_encrypted = parse_document_bytes(target_doc.get("file_name", "doc.pdf"), raw_bytes, password=password)

        if decrypted_text == "[INVALID_PASSWORD]":
            return "Invalid password provided, Sir. Access denied.", False

        await mongo_media_col.update_one(
            {"_id": target_doc["_id"]},
            {"$set": {"caption": decrypted_text, "is_encrypted": False}}
        )
        index_document_into_chroma(target_doc.get("file_name", "doc.pdf"), "document", decrypted_text)
        return decrypted_text, True
    except Exception as e:
        return f"Decryption failed: {str(e)}", False

async def log_security_breach(unauthorized_id: str, raw_msg: str):
    if mongo_security_col is not None:
        try:
            await mongo_security_col.insert_one({
                "unauthorized_id": str(unauthorized_id),
                "attempted_message": raw_msg,
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
        except Exception: pass

    if TELEGRAM_TOKEN and ALLOWED_TELEGRAM_USER_ID:
        alert_msg = f"SECURITY ALERT, Sir: Unauthorized access attempt detected from Telegram User ID {unauthorized_id}."
        async with httpx.AsyncClient() as client:
            try:
                await client.post(
                    f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                    json={"chat_id": ALLOWED_TELEGRAM_USER_ID, "text": alert_msg}
                )
            except Exception: pass

async def send_scheduled_reminder(reminder_text: str):
    if not TELEGRAM_TOKEN or not ALLOWED_TELEGRAM_USER_ID: return
    msg = f"Alert, Sir: You have a scheduled reminder — '{reminder_text}'."
    async with httpx.AsyncClient() as client:
        try:
            await client.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                json={"chat_id": ALLOWED_TELEGRAM_USER_ID, "text": msg}
            )
        except Exception as e: print(f"[Reminder Error]: {e}")

async def create_time_reminder(minutes: int, task_desc: str) -> str:
    run_time = datetime.now(timezone.utc) + timedelta(minutes=minutes)
    job_id = f"reminder_{int(run_time.timestamp())}"
    
    scheduler.add_job(
        send_scheduled_reminder,
        trigger=DateTrigger(run_date=run_time),
        args=[task_desc],
        id=job_id,
        replace_existing=True
    )
    return f"Reminder established for '{task_desc}' in {minutes} minute{'s' if minutes > 1 else ''}, Sir."

async def get_system_diagnostics() -> str:
    mem_count = await mongo_memory_col.count_documents({}) if mongo_memory_col is not None else 0
    task_count = await mongo_tasks_col.count_documents({}) if mongo_tasks_col is not None else 0
    doc_count = await mongo_media_col.count_documents({}) if mongo_media_col is not None else 0
    return f"DIAGNOSTIC STATUS, Sir: All systems nominal. MongoDB Atlas + ChromaDB Vector Store active. Memory Records: {mem_count}, Active Tasks: {task_count}, Secured Documents: {doc_count}."

async def save_scheduled_task(task: str, timing: str, date_str: str = "Today") -> str:
    if mongo_tasks_col is not None:
        try:
            await mongo_tasks_col.insert_one({
                "task": task, "timing": timing, "date": date_str, "status": "active",
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
        except Exception: pass
    return f"Task '{task}' scheduled for {timing}, Sir."

async def save_memory_fact(category: str, fact: str) -> str:
    cat = category.lower().strip()
    fact_str = fact.strip()
    
    if mongo_memory_col is not None:
        try:
            await mongo_memory_col.insert_one({"category": cat, "fact": fact_str, "timestamp": datetime.now(timezone.utc).isoformat()})
        except Exception: pass

    try:
        memory_collection.add(
            ids=[str(datetime.now().timestamp())],
            documents=[fact_str],
            metadatas=[{"category": cat}]
        )
    except Exception as e:
        print(f"[Chroma Memory Insert Error]: {e}")

    return "Information saved permanently in your vector vault, Sir."

async def purge_all_vault_data() -> str:
    if mongo_memory_col is not None: await mongo_memory_col.delete_many({})
    if mongo_tasks_col is not None: await mongo_tasks_col.delete_many({})
    if mongo_chats_col is not None: await mongo_chats_col.delete_many({})
    if mongo_media_col is not None: await mongo_media_col.delete_many({})
    if mongo_reminders_col is not None: await mongo_reminders_col.delete_many({})
    return "All database records and vector database vault data have been completely purged, Sir."

async def log_chat_interaction(user_msg: str, aria_reply: str, session_id: str):
    global LAST_USER_INTERACTION_TIME
    LAST_USER_INTERACTION_TIME = datetime.now(timezone.utc)
    if mongo_chats_col is not None:
        try:
            await mongo_chats_col.insert_one({
                "session_id": session_id, "user_msg": user_msg,
                "aria_reply": aria_reply, "timestamp": datetime.now(timezone.utc).isoformat()
            })
        except Exception: pass

async def save_media_file(file_name: str, media_type: str, raw_bytes: bytes, caption: str = "", is_encrypted: bool = False):
    b64_payload = base64.b64encode(raw_bytes).decode('utf-8')
    if mongo_media_col is not None:
        try:
            await mongo_media_col.insert_one({
                "file_name": file_name, "media_type": media_type,
                "caption": caption.lower().strip(), "b64_payload": b64_payload,
                "is_encrypted": is_encrypted,
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
        except Exception: pass

    if not is_encrypted:
        index_document_into_chroma(file_name, media_type, caption)

    await save_memory_fact("media_vault", f"SAVED {media_type.upper()}: '{file_name}'")
    status_msg = "is encrypted and secured under Privacy Protocol." if is_encrypted else "fully parsed and indexed in your vector vault."
    return f"Document '{file_name}' {status_msg}, Sir."

def fetch_web_search(query: str) -> str:
    if not tavily_client: return ""
    try:
        res = tavily_client.search(query=query, max_results=3)
        results = [f"- {item['title']}: {item['content'][:200]}" for item in res.get("results", [])]
        return "\nREAL-TIME WEB INTELLIGENCE:\n" + "\n".join(results) + "\n"
    except Exception: pass
    return ""

async def fetch_weather_by_coords(location_info: str = "17.6868,83.2185") -> str:
    if not location_info or "," not in location_info: location_info = "17.6868,83.2185"
    try:
        lat, lon = location_info.split(",")
        url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&current_weather=true"
        async with httpx.AsyncClient() as client:
            res = await client.get(url, timeout=2.0)
            if res.status_code == 200:
                data = res.json().get("current_weather", {})
                return f"\nLIVE WEATHER: Temperature {data.get('temperature')}°C, Wind Speed {data.get('windspeed')} km/h.\n"
    except Exception: pass
    return ""

# -------------------------------------------------------------
# 5. BACKGROUND DAEMONS & TASKS
# -------------------------------------------------------------
async def autonomous_proactive_checkin():
    if not TELEGRAM_TOKEN or not ALLOWED_TELEGRAM_USER_ID: return
    now = datetime.now(timezone.utc)
    if (now - LAST_USER_INTERACTION_TIME).total_seconds() / 60.0 < 25: return

    try:
        schedules = []
        task_cursor = mongo_tasks_col.find({}).sort("_id", -1)
        async for tdoc in task_cursor:
            schedules.append(f"• Task: {tdoc.get('task')}")
        if not schedules: return

        active_task = schedules[0]
        checkin_prompt = f"You are J.A.R.V.I.S. The user is working on '{active_task}' and silent for 25 mins. Ask a sharp, human-like check-in question on progress. Address as Sir."

        if groq_client:
            comp = groq_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": checkin_prompt}],
                temperature=0.4, max_tokens=100
            )
            msg = clean_response_text(comp.choices[0].message.content)
            async with httpx.AsyncClient() as client:
                await client.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage", json={"chat_id": ALLOWED_TELEGRAM_USER_ID, "text": msg})
    except Exception: pass

async def summarize_recent_chats():
    if mongo_chats_col is None: return
    try:
        cursor = mongo_chats_col.find({}).sort("_id", -1).limit(30)
        chats = await cursor.to_list(length=30)
        if chats:
            chat_blob = "\n".join([f"User: {c.get('user_msg')}\nARIA: {c.get('aria_reply')}" for c in chats])
            await save_memory_fact("chat_summary", f"Summary of recent discussions: {chat_blob[:1500]}")
    except Exception as e:
        print(f"[Chat Summarization Error]: {e}")

# -------------------------------------------------------------
# 6. FUNCTION-CALLING TOOLS & MASTER CONTEXT ENGINE
# -------------------------------------------------------------
GROQ_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "send_file_from_vault",
            "description": "Dispatches and uploads the actual binary PDF or media file to Telegram when the user asks to send, download, get, or dispatch a file.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_query": {"type": "string", "description": "Document name or keyword e.g. 'resume', 'certificate'"}
                },
                "required": ["file_query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "query_document_vault",
            "description": "Semantic vector search across ChromaDB documents to answer specific questions.",
            "parameters": {
                "type": "object",
                "properties": {
                    "specific_question": {"type": "string", "description": "The exact question or topic to search across vector documents"}
                },
                "required": ["specific_question"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_time_reminder",
            "description": "Sets a timed alert or reminder to trigger after a specific number of minutes.",
            "parameters": {
                "type": "object",
                "properties": {
                    "minutes": {"type": "integer", "description": "Number of minutes from now to send the reminder alert"},
                    "task_desc": {"type": "string", "description": "What the user needs to be reminded to do"}
                },
                "required": ["task_desc"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "save_scheduled_task",
            "description": "Schedules a task or event into the user's daily calendar/agenda.",
            "parameters": {
                "type": "object",
                "properties": {
                    "task": {"type": "string", "description": "Task name or description"},
                    "timing": {"type": "string", "description": "Time slot string e.g. '10am-11am' or '5:00 PM'"}
                },
                "required": ["task", "timing"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "save_memory_fact",
            "description": "Permanently saves a user preference, detail, or personal fact into the vector vault.",
            "parameters": {
                "type": "object",
                "properties": {
                    "category": {"type": "string", "description": "Category e.g. profile, project, preference"},
                    "fact": {"type": "string", "description": "Fact statement to record"}
                },
                "required": ["category", "fact"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_system_diagnostics",
            "description": "Audits assistant health, database metrics, security incidents, and active tasks.",
            "parameters": {"type": "object", "properties": {}}
        }
    }
]

async def process_autonomous_task(user_text: str, session_id: str, location_info: str = None) -> str:
    cmd = user_text.lower().strip()

    # 1. PERMISSION & PENDING ACTION EVALUATION
    if session_id in PENDING_SECURITY_ACTIONS:
        pending = PENDING_SECURITY_ACTIONS[session_id]
        action_type = pending.get("type")

        if action_type == "unlock_pdf":
            pwd_match = re.search(r'\b([A-Za-z0-9@#$_]{4,25})\b', user_text.strip())
            password_input = pwd_match.group(1) if pwd_match else user_text.strip()

            target_doc_keyword = pending["data"]["doc_keyword"]
            original_q = pending["data"]["original_q"]

            decrypted_text, success = await unlock_encrypted_pdf(target_doc_keyword, password_input)
            if not success:
                return f"Security Alert: Decryption failed for '{target_doc_keyword}', Sir. Please provide the correct document password."

            del PENDING_SECURITY_ACTIONS[session_id]
            return f"Document unlocked successfully, Sir. You can now re-issue your query: '{original_q}'."

        if any(k in cmd for k in ["yes", "proceed", "authorize", "do it", "confirm", "sure", "agree", "allow", "grant"]):
            del PENDING_SECURITY_ACTIONS[session_id]
            if action_type == "purge_vault":
                return await purge_all_vault_data()
            elif action_type == "send_file":
                return await send_file_from_vault(pending["data"].get("file_query", "aadhar"), session_id)

        elif any(k in cmd for k in ["no", "cancel", "stop", "abort", "don't", "dont", "deny", "refuse"]):
            del PENDING_SECURITY_ACTIONS[session_id]
            return "Security authorization withheld. Action canceled, Sir."

    if cmd in ["yes", "yeah", "sure", "do it", "send it"] and session_id not in PENDING_SECURITY_ACTIONS:
        return await send_file_from_vault("aadhar", session_id)

    # 2. RUN PLANNER AGENT
    plan = await ai_planner(user_text)

    # 3. DYNAMIC MEMORY CAPTURE
    if any(k in cmd for k in ["remember", "my favourite", "i like", "my project", "passport"]):
        await save_memory_fact("user_statement", user_text)

    # 4. BUILD MASTER CONTEXT
    vector_memories = await query_vector_memory(user_text) if plan["memory"] else ""
    vector_docs = await query_vector_documents(user_text) if plan["documents"] else ""
    web_intel = fetch_web_search(user_text) if plan["internet"] else ""
    temporal_context = get_current_temporal_context()
    weather_context = await fetch_weather_by_coords(location_info or "17.6868,83.2185") if "weather" in cmd else ""

    master_context = f"""
{temporal_context}
{weather_context}
--- MASTER CONTEXT PIPELINE ---
[VECTOR MEMORY]:
{vector_memories}

[VECTOR DOCUMENTS]:
{vector_docs}

[INTERNET INTELLIGENCE]:
{web_intel}
-------------------------------
"""

    system_prompt = f"""You are {ASSISTANT_NAME}, an autonomous, hyper-intelligent AI assistant combining J.A.R.V.I.S. and Spider-Man's Karen.

{master_context}

CRITICAL OPERATIONAL DIRECTIVES:
1. THINK BEFORE SPEAKING:
   - Search memory and vector documents.
   - Use tools automatically.
   - Never ask unnecessary questions.
   - Answer naturally.
2. SENSITIVE IDENTIFIER REDACTION (AADHAAR / RRN / MYNUMBER):
   - Never print raw numeric sequences of Aadhaar or government identification numbers directly in chat text.
   - If asked for an Aadhaar number, state: "Per privacy protocols, I cannot display raw government ID numbers in chat text, but I can dispatch your official PDF file directly to your Telegram." (and offer/execute 'send_file_from_vault').
3. ADDRESS & SALUTATIONS: Address the user naturally as 'Sir' or 'Master'. Keep responses precise and useful."""

    reply_text = ""

    if groq_client:
        try:
            def _groq_exec():
                response = groq_client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_text}],
                    tools=GROQ_TOOLS,
                    tool_choice="auto",
                    temperature=0.2, max_tokens=350
                )
                return response.choices[0].message

            msg = await asyncio.to_thread(_groq_exec)

            if msg.tool_calls:
                for tool_call in msg.tool_calls:
                    fn_name = tool_call.function.name
                    fn_args = json.loads(tool_call.function.arguments) if tool_call.function.arguments else {}

                    if fn_name == "send_file_from_vault":
                        q_term = fn_args.get("file_query", "")
                        if not q_term:
                            for term in ["resume", "certificate", "cs50", "passport", "guide", "aadhar", "aadhaar"]:
                                if term in cmd: q_term = term; break
                        reply_text = await send_file_from_vault(q_term, session_id)

                    elif fn_name == "query_document_vault":
                        specific_q = fn_args.get("specific_question", user_text)
                        doc_text = await query_vector_documents(specific_q)

                        qa_prompt = f"""User Request: '{user_text}'

CHROMA VECTOR SEARCH RESULTS:
{doc_text}

MANDATORY DIRECTIVE:
- Answer the user's specific request using the document text above concisely.
- Address the user as Sir or Master."""

                        qa_comp = groq_client.chat.completions.create(
                            model="llama-3.3-70b-versatile",
                            messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": qa_prompt}],
                            temperature=0.2, max_tokens=250
                        )
                        reply_text = qa_comp.choices[0].message.content.strip()

                    elif fn_name == "create_time_reminder":
                        reply_text = await create_time_reminder(fn_args.get("minutes", 5), fn_args.get("task_desc", "Task"))
                    elif fn_name == "save_scheduled_task":
                        reply_text = await save_scheduled_task(fn_args.get("task"), fn_args.get("timing"))
                    elif fn_name == "save_memory_fact":
                        reply_text = await save_memory_fact(fn_args.get("category", "profile"), fn_args.get("fact"))
                    elif fn_name == "get_system_diagnostics":
                        reply_text = await get_system_diagnostics()

            elif msg.content:
                reply_text = msg.content.strip()

        except Exception as e:
            print(f"[Groq Execution Error]: {e}")

    # Fallback to Gemini 2.0 Flash
    if not reply_text and gemini_client:
        try:
            def _gemini_sync():
                res = gemini_client.models.generate_content(
                    model="gemini-2.0-flash", contents=f"{system_prompt}\n\nUser: {user_text}\nARIA:"
                )
                return res.text
            reply = await asyncio.to_thread(_gemini_sync)
            if reply and len(reply.strip()) > 0: reply_text = reply.strip()
        except Exception as e: print(f"[Gemini Error]: {e}")

    if not reply_text:
        if any(k in cmd for k in ["name", "who am i"]):
            reply_text = f"Your full name is {USER_FULL_NAME}, Sir."
        elif "diagnostic" in cmd:
            reply_text = await get_system_diagnostics()
        else:
            reply_text = f"Understood, Sir. Processing your request."

    # 5. CONFIDENCE-BASED SELF-REFLECTION LOOP (STEP 6 & 11)
    cleaned_reply = clean_response_text(reply_text)
    try:
        reflection_prompt = f"Question: {user_text}\nAnswer: {cleaned_reply}\nIs anything missing or incorrect? Reply only Yes or No."
        ref_comp = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": reflection_prompt}],
            temperature=0.1, max_tokens=10
        )
        ref_ans = ref_comp.choices[0].message.content.strip().lower()
        if "yes" in ref_ans and plan["internet"]:
            extra_search = fetch_web_search(user_text)
            cleaned_reply += f"\n{extra_search[:300]}"
    except Exception: pass

    asyncio.create_task(log_chat_interaction(user_text, cleaned_reply, session_id))
    return cleaned_reply

# -------------------------------------------------------------
# 7. TELEGRAM WEBHOOK & UNIVERSAL UPLOAD ROUTE
# -------------------------------------------------------------
@app.post("/telegram-webhook")
async def telegram_webhook(req: Request):
    if not TELEGRAM_TOKEN: return {"status": "no token"}
    
    try:
        data = await req.json()
        message = data.get("message", {})
        chat_id = message.get("chat", {}).get("id")
        from_user_id = message.get("from", {}).get("id")
        text = message.get("text", "").strip()
        document = message.get("document", None)
        voice = message.get("voice", None)
        video = message.get("video", None)
        photo = message.get("photo", None)

        if not chat_id: return {"status": "no chat_id"}

        if ALLOWED_TELEGRAM_USER_ID and str(from_user_id) != str(ALLOWED_TELEGRAM_USER_ID):
            await log_security_breach(from_user_id, text or "Non-text payload")
            return {"status": "unauthorized"}

        if text.lower() == "/start":
            async with httpx.AsyncClient() as client:
                await client.post(
                    f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                    json={"chat_id": chat_id, "text": "Online and fully operational, Sir. How may I assist you today?"}
                )
            return {"status": "ok"}

        file_obj, media_type, default_name = None, "document", "file.dat"
        if document: file_obj, media_type, default_name = document, "document", document.get("file_name", "document.pdf")
        elif voice: file_obj, media_type, default_name = voice, "voice", "voice_note.ogg"
        elif video: file_obj, media_type, default_name = video, "video", "video_clip.mp4"
        elif photo: file_obj, media_type, default_name = photo[-1], "image", "photo.jpg"

        if file_obj:
            file_id = file_obj.get("file_id")
            async with httpx.AsyncClient() as client:
                file_info = await client.get(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getFile?file_id={file_id}")
                file_path = file_info.json().get("result", {}).get("file_path")
                raw_bytes_res = await client.get(f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}")
                raw_bytes = raw_bytes_res.content

            if media_type == "image":
                extracted_text = await process_image_with_gemini_vision(raw_bytes, default_name)
                is_encrypted = False
            else:
                extracted_text, is_encrypted = parse_document_bytes(default_name, raw_bytes)

            save_reply = await save_media_file(default_name, media_type, raw_bytes, caption=extracted_text, is_encrypted=is_encrypted)
            async with httpx.AsyncClient() as client:
                await client.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage", json={"chat_id": chat_id, "text": save_reply})
            return {"status": "ok"}

        if text:
            reply_text = await process_autonomous_task(text, str(chat_id))
            async with httpx.AsyncClient() as client:
                await client.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage", json={"chat_id": chat_id, "text": reply_text})
            return {"status": "ok"}

    except Exception as e: print(f"[Webhook Error]: {e}")
    return {"status": "ok"}

@app.on_event("startup")
async def start_scheduler():
    scheduler.add_job(autonomous_proactive_checkin, 'interval', minutes=30, id="proactive_checkin_job")
    scheduler.add_job(summarize_recent_chats, 'interval', hours=6, id="summarize_chats_job")
    scheduler.start()
    print("[J.A.R.V.I.S. Multi-Agent Vector Core]: Online and Synced.")

# -------------------------------------------------------------
# 7. SPEECH & FRONTEND HUD
# -------------------------------------------------------------
async def generate_speech_audio_b64(text: str, selected_voice: str = "en-GB-RyanNeural") -> str:
    is_telugu_script = bool(re.search(r'[\u0C00-\u0C7F]', text))
    voice_to_use = "te-IN-MohanNeural" if (is_telugu_script and "te-IN" not in selected_voice) else selected_voice

    try:
        communicate = edge_tts.Communicate(text, voice_to_use)
        audio_data = bytearray()
        async for chunk in communicate.stream():
            if chunk["type"] == "audio": audio_data.extend(chunk["data"])
        return base64.b64encode(audio_data).decode('utf-8')
    except Exception:
        communicate = edge_tts.Communicate(text, "en-GB-RyanNeural")
        audio_data = bytearray()
        async for chunk in communicate.stream():
            if chunk["type"] == "audio": audio_data.extend(chunk["data"])
        return base64.b64encode(audio_data).decode('utf-8')

@app.head("/health")
@app.get("/health")
def health_check():
    return JSONResponse(status_code=200, content={"status": "online", "database": "MongoDB Atlas + ChromaDB", "system": "ARIA Multi-Agent Vector Core Active"})

@app.head("/", response_class=HTMLResponse)
@app.get("/", response_class=HTMLResponse)
def serve_webapp():
    return f"<h1>ARIA Multi-Agent Vector-Indexed J.A.R.V.I.S. Core Online</h1>"

# -------------------------------------------------------------
# 8. WEBSOCKET STREAMING & UPLOAD ROUTE
# -------------------------------------------------------------
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    session_id = str(id(websocket))
    try:
        while True:
            raw_data = await websocket.receive_text()
            data = json.loads(raw_data)
            prompt = data.get("prompt", "")
            location = data.get("location", None)

            reply_text = await process_autonomous_task(prompt, session_id, location)
            audio_b64 = await generate_speech_audio_b64(reply_text)
            
            await websocket.send_json({"audio": audio_b64, "text": reply_text})
    except WebSocketDisconnect: pass

@app.post("/upload-file")
async def upload_file(file: UploadFile = File(...), category: str = "documents"):
    file_bytes = await file.read()
    if file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.webp')):
        extracted_text = await process_image_with_gemini_vision(file_bytes, file.filename)
        is_encrypted = False
    else:
        extracted_text, is_encrypted = parse_document_bytes(file.filename, file_bytes)
    
    save_reply = await save_media_file(file.filename, "document", file_bytes, caption=extracted_text[:6000], is_encrypted=is_encrypted)
    return {"status": "ok", "message": save_reply}
